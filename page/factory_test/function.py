#standard library
import sys
import os
import time
from enum import Enum
import csv
import serial
import re
from datetime import datetime
sys.path.append(os.getcwd())

#third party
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import openpyxl
from openpyxl.styles import PatternFill

#module
from interface.interface_worker import  GenericWorker, do_in_thread
from interface.bluetooth.bluetooth import Bluetooth_LE, dongle
from common.convert import Data_convertor as cov
from common.config import Config_creator
from page.factory_test.CR1_bt import test, device, fw_update
from page.factory_test.err_cls import *
from common.UI_object import Timer_generator
import logging
logger = logging.getLogger('factory_test')

#sub-module
if __name__ == '__main__':
    from UI_factory_test_page import Ui_factory_test_func
else:
    from .UI_factory_test_page import Ui_factory_test_func

cov = cov()
worker = GenericWorker()
do_in_thread = do_in_thread(worker)

RSSI_LIMIT  = -60   # dBm, device will only display when the RSSI is higher than this value
SCAN_TIME   =   3   # seconds

class factory_test_func(QWidget, Ui_factory_test_func):
    def __init__(self):
        ## Set custom value ##
        self.dongle_used = dongle.bleuio.value # current use dongle module

        ## Init value ##
        super(factory_test_func, self).__init__()
        self.setupUi(self)
        self.config = Config_creator('factory_test_func')
        self.config_path = self.config.get_path()
        self._ble = worker.ble
        if self._ble.client == None:
            self._ble.select_interface(self.dongle_used)
        self._dev_common = device()
        self._display_devices = []
        self._barometer = None
        self._is_connecting = False # flag for connecting, control Btn Behavior
        self._gui_mes_sn = 0
        self._fw_version = 'N/A'
        self._fw_build_time = 'N/A'
        self._timer = {
            'connChecker': {'timer': QTimer(), 'interval': 5000, 'functions': [self._check_conn_status]},
            'awakener': {'timer': QTimer(), 'interval': 20000, 'functions': [self._dev_common.check_communicable]},
        }
        Timer_generator(self, self._timer)

        worker.message.connect(self._show_GUI_message)
        # self.LED_GuiNormal.set_color(2)
        self.Btn_scanDevice.clicked.connect(self._handle_scan_Btn_event)
        self.Btn_connect.clicked.connect(self.handle_connect_Btn_event)
        self.Btn_devTest.clicked.connect(self.handle_test_Btn_event)
        self.Btn_selectFile.clicked.connect(self._select_file)
        self.Btn_updataFw.clicked.connect(self.handle_updateFw_Btn_event)
        self.Btn_COMportOpen.clicked.connect(self.handle_COM_open_Btn_event)
        self.Btn_Calibrate.clicked.connect(self.handle_calPSensor_Btn_event)
        self.Btn_findWatch.clicked.connect(self._handle_find_watch_event)

    @do_in_thread
    def _handle_scan_Btn_event(self):
        try:
            # UI init
            self._set_Btn_disable(True)
            self._show_GUI_message('SCAN START')
            worker.UI.emit(lambda: self.Btn_scanDevice.setText("WAIT  5  SEC"))
 
            # action
            self._ble.scan(timeout=SCAN_TIME, scan_cb=self._store_scan_data)
            self._set_devices_list()

            #UI deinit
            self._set_Btn_disable(False)
            self._show_GUI_message('SCAN END')
            worker.UI.emit(lambda: self.Btn_scanDevice.setText("SCAN  DEVICE"))
        except Exception as e:
            self._handle_err(e)

    def _set_devices_list(self):
        worker.UI.emit(lambda: self._remove_devices_in_listWidget())
        self._sort_devices_by_rssi()
        worker.UI.emit(lambda: self.add_devices_in_listWidget())

    def _store_scan_data(self, msg: dict, a, b):
        self._display_devices = []
        for value in msg.values():
            device_name, address, rssi, packet_data = self._get_inf_from_dongle_data(value)
            if rssi > RSSI_LIMIT:
                self._display_devices.append(
                    bt_dev_display_info(address, device_name, rssi, packet_data)
                )

    def _sort_devices_by_rssi(self):
        """
        @brief: Sort the devices by rssi from high to low.
        @note: Algorithm: merge sort.
        """      
        def merge(left, right):
            result = []

            while len(left) and len(right):
                if (left[0].rssi < right[0].rssi):
                    result.append(left.pop(0))
                else:
                    result.append(right.pop(0))

            result = result+left if len(left) else result+right
            return result

        def mergeSort(array):
            if len(array) < 2:
                return array

            mid = len(array)//2
            leftArray = array[:mid]
            rightArray = array[mid:]

            return merge(mergeSort(leftArray),mergeSort(rightArray))
        
        self._display_devices = mergeSort(self._display_devices)
        self._display_devices.reverse()
    
    def _remove_devices_in_listWidget(self, how_many=0):
        """
        @brief: Remove removes the item from the top of the list.
        @param: how_many(int), if 0 means will remove all.
        """
        if how_many == 0:
            self.listWidget_deviceList.clear()
        else:
            for i in range(how_many):
                self.listWidget_deviceList.removeItemWidget(
                    self.listWidget_deviceList.takeItem(0)
                )

    def add_devices_in_listWidget(self):
        font = QFont()
        font.setFamily("Consolas")
        font.setPointSize(16)
        
        for dev in self._display_devices:
            new_item = QListWidgetItem()
            new_item.setText(dev.widgetItemText)
            new_item.setFont(font)
            new_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.listWidget_deviceList.addItem(new_item)

    def _get_inf_from_dongle_data(self, value):
        """
        @return (name, address, rssi, packet_data)
        """
        rtn = ()
        if self.dongle_used == dongle.bleuio.value:
            rtn = (
                value['name'], value['address'], value['rssi'], value['pkt_dat']
            )
        elif self.dongle_used == dongle.blueGiga.value:
            rtn = (
                value.name, value.address, value.rssi, value.packet_data
            )
        return rtn
    
    @do_in_thread
    def _connect_device(self):
        try:
            item_select = self.listWidget_deviceList.item(self.listWidget_deviceList.currentRow())
            if item_select == None:
                raise conditionShort('NO SELECT DEV')

            # UI init
            self._show_GUI_message('CONN START')
            worker.UI.emit(lambda: self.Btn_connect.setText("CONNECTING ..."))

            # action
            if item_select.text().split(' - ')[0] != 'CREST-CR1': 
                raise conditionShort('Select device not CR1')
            MAC_addrs = item_select.text().split(' - ')[-1].split(']')[-1]
            self._ble.connect(MAC_addrs)
            time.sleep(2)
            self._is_connecting = self._ble.is_connect()
        finally:
            if self._is_connecting:
                self._dev_common.subscribe_app_uuid()
                worker.UI.emit(lambda: self.label_currDevAddr.setText(MAC_addrs))
                self._fw_version = self._dev_common.get_fw_ver()
                self._fw_build_time = self._dev_common.get_build_time()
                worker.UI.emit(lambda: self.label_currFwVer.setText(self._fw_version))
                worker.UI.emit(lambda: self.label_currFwTime.setText(self._fw_build_time))
                self._start_bt_timer()
                worker.UI.emit(lambda: self.Btn_connect.setText("<< DISCONNECT"))
            else:
                worker.UI.emit(lambda: self.Btn_connect.setText( "CONNECT >>"))
            self._show_GUI_message('CONN END' if self._is_connecting else 'CONN FAIL')

    @do_in_thread
    def _disconnect_dev(self):
        self._ble.disconnect()
        time.sleep(0.5)
        self._is_connecting = self._ble.is_connect()
        if not self._is_connecting:
            self._fw_build_time = 'N/A'
            self._fw_version = 'N/A'
            self._stop_bt_timer()
            worker.UI.emit(lambda: self.label_currDevAddr.setText('N/A'))
            worker.UI.emit(lambda: self.label_currFwVer.setText(self._fw_version))
            worker.UI.emit(lambda: self.label_currFwTime.setText(self._fw_build_time))
            self._show_GUI_message('DEV DISCONN')
            worker.UI.emit(lambda: self.Btn_connect.setText( "CONNECT >>"))

    @do_in_thread
    def handle_connect_Btn_event(self):
        try:
            # UI init
            self._set_Btn_disable(True)

            # action
            self._disconnect_dev() if self._is_connecting else self._connect_device()
        except Exception as e:
            self._handle_err(e)
        finally: # UI deinit
            self._set_Btn_disable(False)

    @do_in_thread
    def handle_test_Btn_event(self):
        try:
            self._stop_bt_timer()
            if self._is_connecting == False: 
                raise serverNoConnect()
            if self.lineEdit_DevSN.text() == '': 
                raise conditionShort('NO SN')
            self._set_Btn_disable(True)
            if not self._dev_common.check_communicable(): 
                raise serverNoResponse('TEST FAIL')
            
            # UI init
            self.lineEdit_DevSN.setReadOnly(True) # Avoid changing the SN during the test to cause the wrong file name of the test result
            worker.UI.emit(lambda: self.Btn_devTest.setText('TESTING ...'))

            # action
            self._file_name = self.lineEdit_DevSN.text()
            self._dev_common.set_sn(self._file_name)
            objTemp = test()
            if objTemp.start(): # if test is finished
                self._file_data = objTemp.get_result_data()
                self._generate_csv_file()
                self._add_test_history()
                self._show_test_result()
        except Exception as e:
            self._handle_err(e)
        else:
            self._show_GUI_message('TEST END')
        finally: # UI deinit
            self._start_bt_timer()
            self.lineEdit_DevSN.setReadOnly(False) # Unlock SN read-only restrictions
            self._set_Btn_disable(False)
            worker.UI.emit(lambda: self.Btn_devTest.setText("<<  COMPREHENSIVE TEST  >>"))

    def _generate_csv_file(self):
        self._isAllPass = True
        # folder = r'.\page\factory_test\test_report\\'
        folder = r'.\test_report\\'
        with open(folder + self._file_name + '.csv', 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows([
                ['Firmware version:', self._fw_version],
                ['Firmware build time:', self._fw_build_time],
                [' '],
                ['Test case', 'Value/Error code', 'Note']
            ])
            for test_case, value in self._file_data.items():
                writer.writerow([test_case, value[0], value[1]])
                if value[1] != 'Pass':
                    self._isAllPass = False
    
    def _generate_xslx_file(self):
        self._isAllPass = True
        wb = openpyxl.Workbook()
        s1 = wb[self._file_name]
        s1['A1'].value = 'Firmware version:'
        s1['A2'].value = self._fw_version
        s1['B1'].value = 'Firmware build time:'
        s1['B2'].value = self._fw_build_time
        s1['D1'].value = 'Test case'
        s1['D2'].value = 'Value/Error code'
        s1['D3'].value = 'Note' 
        row = 4
        for test_case, value in self._file_data.items():
            row += 1
            s1.cell(row, 0).value = test_case
            s1.cell(row, 1).value = value[0]
            s1.cell(row, 2).value = value[1]
            if value[1] != 'Pass':
                self._isAllPass = False
                for column in range(3):
                    s1.cell(row, column).fill = PatternFill(
                        fill_type='solid', fgColor='FF0000'
                    )
        wb.save(self._file_name + '.xlsx')
    
    def _show_test_result(self):
        # folder = r'.\page\factory_test\test_report\\'
        folder = r'.\test_report\\'
        with open(folder + self._file_name + '.csv', 'r') as csvfile:
            r = csv.reader(csvfile)
            for row in list(r):
                msg = ''
                for data in row:
                    msg += data + ', '
                self._show_GUI_message(msg[:-2])

    def _add_test_history(self):
        self.textBrowser_testHistory.append(
            datetime.now().strftime("%m-%d %H:%M") 
            + '    '
            + (' ' if self._isAllPass else '*') 
            + self._file_name + '.csv'
            + '\n'
        )

    def _select_file(self):
        filename, filetype = QFileDialog.getOpenFileName(self, "Open file", "./") # start path
        logger.info('Select file: ' + filename + '.' + filetype)
        worker.UI.emit(lambda: self.textEdit_filePath.setText(filename))
        self.textEdit_filePath.setFontFamily("Consolas")
        self.textEdit_filePath.setFontPointSize(16)

    @do_in_thread
    def handle_updateFw_Btn_event(self):
        try:
            if self._is_connecting == False:
                raise serverNoConnect()
            if self.textEdit_filePath.toPlainText() == '':
                raise conditionShort('NO FILE')
            self._set_Btn_disable(True)
            if not self._dev_common.check_communicable():
                raise serverNoResponse('TEST FAIL')

            # UI init
            self._show_GUI_message('START UPDATE FW')
            worker.UI.emit(lambda: self.Btn_updataFw.setText('UPDATING ...'))

            # action
            self._stop_bt_timer()
            self._dev_common.subscribe_ota_uuid()
            ota_update = fw_update(self.textEdit_filePath.toPlainText())
            error_code =  ota_update.start()
            if error_code is not None: # end of ota update
                self._show_GUI_message(str(error_code))
                self._disconnect_dev()
        except Exception as e:
            self._handle_err(e)
        finally: # UI deinit
            self._start_bt_timer()
            self._set_Btn_disable(False)
            worker.UI.emit(lambda: self.Btn_updataFw.setText('Finish Update'))
            time.sleep(1.5)
            worker.UI.emit(lambda: self.Btn_updataFw.setText('Update FW'))

    def handle_COM_open_Btn_event(self):
        try:
            port_num = self.lineEdit_COMport.text()
            if port_num == '':
                raise conditionShort('NO COM')

            self.lineEdit_COMport.setReadOnly(True)
            comport = 'COM' + port_num
            if self._barometer == None or self._barometer.portstr != comport:
                self._barometer = serial.Serial(comport)
        except IOError:
            logger.error('Cannot open ' + comport)
            self._show_GUI_message('COM PORT ERR')
        except Exception as e:
            self._handle_err(e)
        else:
            self._show_GUI_message('COM PORT OK')
        finally:
            self.lineEdit_COMport.setReadOnly(False)

    @do_in_thread
    def handle_calPSensor_Btn_event(self):
        try:
            self._stop_bt_timer()
            if self._barometer == None:
                raise conditionShort('NO COM')
            if not self._is_connecting:
                raise serverNoConnect()
            self._set_Btn_disable(True)
            if not self._dev_common.check_communicable():
                raise serverNoResponse('TEST FAIL')
            
            # UI init
            worker.UI.emit(lambda: self.Btn_Calibrate.setText('CALIBRATING ...'))
            
            # action
            data_p, data_t= self._get_barometer_data()
            if 900 < data_p and data_p < 1100:
                if self._dev_common.calibrate_pSensor(data_p):
                    self._show_GUI_message('CALIBRATE PASS')
                else:
                    self._show_GUI_message('CALIBRATE FAIL')
            else:
                self._show_GUI_message('Pressure Invalid')
        except Exception as e:
            self._handle_err(e)
        finally: # UI deinit
            self._start_bt_timer()
            self._set_Btn_disable(False)
            worker.UI.emit(lambda: self.Btn_Calibrate.setText(' Calibrate >>'))
    
    def _get_barometer_data(self):
        # sync data frame
        pinst_read = self._barometer.read(1) # read start word
        while(list(pinst_read)[0] != 2):
            pinst_read = self._barometer.read(1) # find start word
        pinst_read = self._barometer.read(15) # discard this data

        # find pressure and temperature data
        atm_valid = False
        while self._barometer.in_waiting >= 16 or atm_valid == False:
            pinst_read = self._barometer.read(1) # read start word
            pinst_read = self._barometer.read(14) # pinst datasize = 14
            logger.debug('pinst_read: ' + pinst_read.decode('utf8'))
            if re.findall('420101', pinst_read.decode('utf8')) != []:
                temperature = float(re.findall('420101(\d+)', pinst_read.decode('utf8'))[0])/10
                logger.debug(f'Instrument Temperature: {temperature}')
            if re.findall('439101', pinst_read.decode('utf8')) != []:
                atm_pressure = float(re.findall('439101(\d+)', pinst_read.decode('utf8'))[0])/10
                self._show_GUI_message(f'Instrument Pressure: {atm_pressure}')
                logger.debug(f'Instrument Pressure: {atm_pressure}')
                atm_valid = True
            pinst_read = self._barometer.read(1) # read end word
        return (atm_pressure, temperature)
    
    @do_in_thread
    def _check_conn_status(self):
        if not self._ble.is_connect():
            self._disconnect_dev()
    
    def _handle_err(self, err: Exception):
        err_type = type(err)
        if serverNoResponse == err_type:
            self._show_GUI_message(err.msg)
            self._disconnect_dev()
            self._set_Btn_disable(False)
        elif serverNoConnect == err_type:
            self._show_GUI_message('NO CONN')
        elif conditionShort == err_type:
            self._show_GUI_message(err.msg)
        else:
            # self.LED_GuiNormal.set_color(1)
            logger.error(err)
            self._show_GUI_message('UNKNOW ERR')

    @do_in_thread  
    def _set_Btn_disable(self, is_disable: bool):
        self.Btn_scanDevice.setDisabled(is_disable)
        self.Btn_connect.setDisabled(is_disable)
        self.Btn_devTest.setDisabled(is_disable)
        self.Btn_updataFw.setDisabled(is_disable)
        self.Btn_Calibrate.setDisabled(is_disable)
        self.Btn_findWatch.setDisabled(is_disable)
    
    def _show_GUI_message(self, info: str):
        self._gui_mes_sn += 1
        mbox_disp_inf = {
            'NO SELECT DEV': 'Please select a device, before connect.',
            'NO SN': 'Please enter the serial number of the device, before test begin.',
            'IN TESTING': 'The device is being tested, please wait for the test to finish and try again',
            'NO CONN': 'Please connect the device, and try again.',
            'START UPDATE FW': 'Start update firmware, please wait...',
            'FW UPDATE SUCC': 'FW update successfully.',
            'OPEN FILE ERR': 'Open file failed.',
            'FW HEADER ERR': 'Check firmware update file is effective.',
            'FW DATA ERR': 'Incomplete file, check firmware update file is effective.',
            'VERIFY FW ERR': 'firmware file has been sent, but the verifiy failed, please send the update file again',
            'COM PORT ERR': 'Cannot open COM port, please check the COM port is correct.',
            'COM PORT OK': 'COM port has been opened successly.',
            'TEST FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST1 FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST2 FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST3 FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST4 FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST5 FAIL': 'Bluetooth device no response, please connect the device again.',
            'SCAN START': 'Scanning for bleutooth devices, please wait...',
            'SCAN END': 'Scan finished.',
            'CONN START': 'Connecting to device, please wait...',
            'CONN END': 'Connect success.',
            'CONN FAIL': 'Connect failed, device may not trun on.',
            'CALIBRATE PASS': 'Calibrate barometric success.',
            'RECIVE FAIL': 'Bluetooth device no response',
            'Pressure Invalid': 'Instrument Pressure out of range',
            'TEST END': 'Test finished.',
            'DEV DISCONN': 'Bluetooth device disconnected.',
            'NO FILE': 'Please select a file, before update firmware.',
            'NO COM': 'Please open a COM port, and try again.',
            'UNKNOW ERR': 'Please turn off and remove the bluetooth receiver then plug it back in and turn on the GUI',
            'CALIBRATE FAIL': 'Calibrate barometric failed, device recive error.',
        }
        now = datetime.now()
        time = now.strftime("%Y-%m-%d %H:%M:%S")
        if info in mbox_disp_inf.keys():
            message = f'[{self._gui_mes_sn}]{time:<25}{info:<20}- {mbox_disp_inf[info]}'
        else:
            message = f'[{self._gui_mes_sn}]{time:<25}{info}'
        worker.UI.emit(lambda: self.textBrowser_guiMessage.append(message))

    @do_in_thread
    def _start_bt_timer(self):
        worker.UI.emit(lambda: self._timer['connChecker']['timer'].start(5000)) # ms
        worker.UI.emit(lambda: self._timer['awakener']['timer'].start(20000)) # ms
    
    @do_in_thread
    def _stop_bt_timer(self):
        worker.UI.emit(lambda: self._timer['connChecker']['timer'].stop())
        worker.UI.emit(lambda: self._timer['awakener']['timer'].stop())

    @do_in_thread
    def _handle_find_watch_event(self):
        try:
            if not self._is_connecting:
                raise serverNoConnect()
            if len(self._dev_common.find_watch()) == 0:
                raise serverNoResponse('TEST FAIL')
        except Exception as e:
            self._handle_err(e)
        

class bt_dev_display_info(object):
    def __init__(self, ble_addr: str, device_name: str, rssi: int, packet_data):
        self.addr = ble_addr
        self.device_name = device_name
        self.rssi = rssi
        self.packet_data = packet_data
        self.widgetItemText = f'{device_name} - ({rssi}) - {ble_addr}'

if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = factory_test_func()
    win.show()
    sys.exit(app.exec_())
