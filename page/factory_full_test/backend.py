#standard library
import sys
import os
import time
import serial
import re
import csv
sys.path.append(os.getcwd())

#third party
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import openpyxl
from openpyxl.styles import PatternFill

#module
from pygatt.backends import BLEAddressType
from interface.interface_worker import  GenericWorker, do_in_thread
from page.factory_full_test.msg import NotifyType
from interface.bluetooth.bluetooth import dongle
from common.config import Config_creator
from Script.Bluetooth.CR1_bt import test, device, fw_update, log_readout, calibrate_pressure
from Error.err_cls import *
from common.UI_object import Timer_generator
import logging
logger = logging.getLogger('rotatingFileLogger')

worker = GenericWorker()
do_in_thread = do_in_thread(worker)

class backend():
    __instance = None
    def __new__(cls):
        if cls.__instance == None:
            cls.__instance= super().__new__(cls)
        return cls.__instance

    def __init__(self):
        self.config = Config_creator('factory_test_func')
        self.config.add_config('factory_test_func', 'dongle', dongle.bleuio.value)
        self.config.add_config('factory_test_func', 'RSSI_LIMIT', -100)  # dBm, device will only display when the RSSI is higher than this value
        self.config.add_config('factory_test_func', 'sacn_time', 3)     # seconds
        self.config.add_config('factory_test_func', 'test_device', 'CREST-CR1')

        self._dongle_current = self.config.get_config()['factory_test_func']['dongle']
        self._scan_time = self.config.get_config()['factory_test_func']['sacn_time']
        self._test_device = self.config.get_config()['factory_test_func']['test_device']
        self._rssi_limit = self.config.get_config()['factory_test_func']['RSSI_LIMIT']
        self._display_devices = []
        self._test_data = {}
        self._barometer = None
        self._is_barometer_open = False
        self._instrument_p = 0.0    # unit: hPa
        self._instrument_t = 0.0    # unit: degree C
        self._is_barometer_refrashing = False
        self._fw_ver = ''
        self._fw_build_t = ''
        self._battery_volt = ''
        self._dev_common = device()
        self._msg_content = {}
        self._timer = {
            'bt5Sec': {
                'timer': QTimer(), 
                'interval': 5000, 
                'functions': [
                    self._check_conn_status,
                    lambda: worker.message.emit([NotifyType.ACTION.value,'updateBattVolt']),
                ]
            },
            'bt10Sec': {
                'timer': QTimer(), 
                'interval': 10000, 
                'functions': [
                    lambda: worker.message.emit([NotifyType.ACTION.value,'updateBattVolt'])
                ]
            },
            'reflashCalPValue': {
                'timer': QTimer(), 
                'interval': 6500, 
                'functions': [self._get_barometer_data]
            },
        }
        Timer_generator(self, self._timer)
        self._ble = worker.ble
        if self._ble.client == None:
            self._ble.select_interface(self._dongle_current)
        if self._ble.is_connect():
            self._ble.disconnect()

    def do_scan(self):
        """
        @brief: Do scan btn event things.
        """
        worker.message.emit([NotifyType.GUI_LOG.value, 'SCAN START'])
        self._ble.scan(
            timeout=self._scan_time, 
            scan_cb=self._store_scan_data
        )
        self._sort_devices_by_rssi()
        worker.message.emit([NotifyType.GUI_LOG.value, 'SCAN END'])

    def get_scan_data(self) -> list:
        return self._display_devices

    def do_conn_dev(self, addr, addr_type, dev_name) -> bool:
        """
        @brief: Do connect btn event things.
        @return Is now connecting?
        """
        self._ble.connect(
            addr, 
            BLEAddressType.public if addr_type == '[0' else BLEAddressType.random
        )
        is_conn = self._ble.is_connect()
        if is_conn and self._test_device == dev_name:
            self._dev_common.subscribe_app_uuid()
            self._fw_ver = self._dev_common.get_fw_ver()
            self._fw_build_t = self._dev_common.get_build_time()
            self._battery_volt = self._dev_common.get_battery_volt()
            self._dev_common.find_watch()
            self._start_bt_timer()
        return is_conn
    
    def get_fw_version(self) -> str:
        return self._fw_ver
    
    def get_fw_build_time(self) -> str:
        return self._fw_build_t
    
    @do_in_thread
    def get_battery_volt(self):
        return self._battery_volt

    def do_disconn_dev(self) -> bool:
        """
        @brief: Do disconnect things.
        @return: Is now connecting?
        """
        self._stop_bt_timer()
        self._ble.disconnect()
        time.sleep(0.3)
        is_conn = self._ble.is_connect()
        if not is_conn:
            worker.message.emit([NotifyType.GUI_LOG.value, 'DEV DISCONN'])
        else:
            self._start_bt_timer()
        return is_conn
    
    def do_open_com_port(self, port_num):
        comport = 'COM' + port_num
        if self._barometer == None or self._barometer.portstr != comport:
            self._barometer = serial.Serial(comport)
            self._get_barometer_data()
            self._barometer.timeout = 6.0
            self._timer['reflashCalPValue']['timer'].start()
    
    @do_in_thread
    def do_cal_pSensor(self):
        try:
            is_succ, before_p, after_p, offset = False, 0, 0, 0
            self._stop_bt_timer()
            if self._barometer == None:
                raise conditionShort('NO COM')
            if not self._dev_common.check_communicable():
                raise serverNoResponse('RECIVE FAIL')
            
            worker.message.emit([
                NotifyType.GUI_LOG.value, f'Instrument Pressure = {self._instrument_p}'
            ])
            if (900 < self._instrument_p 
                and self._instrument_p < 1100 
                and not self._is_barometer_refrashing):
                objTemp = calibrate_pressure(self._instrument_p)
                before_p, after_p, offset = objTemp.start()
                is_succ = (round(after_p) == round(self._instrument_p))
            else:
                worker.message.emit([NotifyType.GUI_LOG.value, 'Pressure Invalid'])
        finally:
            self._start_bt_timer()
            return is_succ, before_p, after_p, offset

    @do_in_thread
    def _get_barometer_data(self):
        try:
            # sync data frame
            pinst_read = self._read_baro_pinst(1) # read start word
            while(list(pinst_read)[0] != 2):
                pinst_read = self._read_baro_pinst(1) # find start word
            pinst_read = self._read_baro_pinst(15) # discard this data

            # find pressure and temperature data
            atm_valid = False
            while self._barometer.in_waiting >= 16 or atm_valid == False:
                self._is_barometer_refrashing = True
                pinst_read = self._read_baro_pinst(1) # read start word
                pinst_read = self._read_baro_pinst(14) # pinst datasize = 14
                if re.findall('420101', pinst_read.decode('utf8')) != []:
                    self._instrument_t = float(re.findall('420101(\d+)', pinst_read.decode('utf8'))[0])/10
                    logger.debug(f'Instrument Temperature: {self._instrument_t}')
                if re.findall('439101', pinst_read.decode('utf8')) != []:
                    self._instrument_p = float(re.findall('439101(\d+)', pinst_read.decode('utf8'))[0])/10
                    logger.debug(f'Instrument Pressure: {self._instrument_p}')
                    atm_valid = True
                pinst_read = self._read_baro_pinst(1) # read end word
                self._is_barometer_refrashing = False
        except Exception as e:
            logger.error(e)
    
    def _read_baro_pinst(self, size):
        data = self._barometer.read(size)
        if len(data) == size and self._is_barometer_open == True: # normal
            return data
        elif len(data) < size and self._is_barometer_open == False: # error
            raise Exception('Barometer COM Port is closed')
        elif len(data) == size and self._is_barometer_open == False: # find normal
            worker.message.emit([NotifyType.ACTION.value, 'comPortStatusNormal'])
            self._barometer.timeout = 5.0
            self._is_barometer_open = True
            worker.message.emit([NotifyType.GUI_LOG.value, 'COM PORT OK'])
            return data
        elif len(data) < size and self._is_barometer_open == True:  # find error
            worker.message.emit([NotifyType.ACTION.value, 'comPortStatusError'])
            self._barometer.timeout = 1.0
            self._is_barometer_open = False
            worker.message.emit([NotifyType.GUI_LOG.value, 'COM PORT ERR'])
            raise Exception('Barometer COM Port is closed')

    def _start_bt_timer(self):
        worker.UI.emit(lambda: self._timer['bt5Sec']['timer'].start(5000)) # ms
        worker.UI.emit(lambda: self._timer['bt10Sec']['timer'].start(10000)) # ms

    def _stop_bt_timer(self):
        worker.UI.emit(lambda: self._timer['bt5Sec']['timer'].stop())
        worker.UI.emit(lambda: self._timer['bt10Sec']['timer'].stop())
        time.sleep(1)

    def _store_scan_data(self, msg: dict, a, b):
        self._display_devices = []
        for value in msg.values():
            device_name, address, rssi, packet_data = self._get_inf_from_dongle_data(value)
            if (rssi > self._rssi_limit and device_name == self._test_device):
                self._display_devices.append(
                    bt_dev_display_info(address, device_name, rssi, packet_data)
                )
    
    def _sort_devices_by_rssi(self):
        """
        @brief: Sort the devices by rssi from high to low.
        @note: Algorithm: merge sort.
        """      
        def merge(left, right):
            result = []

            while len(left) and len(right):
                if (left[0].rssi < right[0].rssi):
                    result.append(left.pop(0))
                else:
                    result.append(right.pop(0))

            result = result+left if len(left) else result+right
            return result

        def mergeSort(array):
            if len(array) < 2:
                return array

            mid = len(array)//2
            leftArray = array[:mid]
            rightArray = array[mid:]

            return merge(mergeSort(leftArray),mergeSort(rightArray))
        
        self._display_devices = mergeSort(self._display_devices)
        self._display_devices.reverse()
    
    def _get_inf_from_dongle_data(self, value):
        """
        @return (name, address, rssi, packet_data)
        """
        rtn = ()
        if self._dongle_current == dongle.bleuio.value:
            rtn = (
                value['name'], value['address'], value['rssi'], value['pkt_dat']
            )
        elif self._dongle_current == dongle.blueGiga.value:
            rtn = (
                value.name, value.address, value.rssi, value.packet_data
            )
        return rtn
    
    def do_test(self):
        """
        @brief: Do test btn event things.
        @return: if test finished return True, else throw a Exception.
        """
        try:
            self._stop_bt_timer()
            self._test_data = {}
            is_finish = False
            if not self._dev_common.check_communicable():
                raise serverNoResponse('RECIVE FAIL')
            objTemp = test()
            if objTemp.start(): # if test finished
                self._test_data = objTemp.get_result_data()
                is_finish = True
        finally:
            self._start_bt_timer()
            return is_finish
    
    def get_test_data(self):
        """
        @brief: Get test data.
        @return: (test_data, is_all_pass)
        """
        return (self._test_data, self._is_test_all_pass())

    def _is_test_all_pass(self):
        is_all_pass = True
        for test_case, value in self._test_data.items():
            if value[1] != 'Pass':
                is_all_pass = False
                break
        return is_all_pass
    
    def _check_conn_status(self):
        if not self._ble.is_connect():
            self._stop_bt_timer()
            worker.message.emit([NotifyType.ACTION.value, 'Device been disconnected'])
    
    def do_update_fw(self, file_path: str):
        """
        @brief: Do update fw btn event things.
        """
        is_finish = False
        is_succ = False
        if not self._dev_common.check_communicable():
            raise serverNoResponse('RECIVE FAIL')
        self._stop_bt_timer()
        self._dev_common.subscribe_ota_uuid()
        ota_update = fw_update(file_path)
        error_code =  ota_update.start()
        if error_code is not None: # end of ota update
            worker.message.emit([NotifyType.GUI_LOG.value, str(error_code)])
            is_finish = True
            is_succ = 'FW UPDATE SUCC' == error_code
        self._start_bt_timer()
        return is_finish, is_succ

    def find_watch(self):
        """
        @brief: Do Find watch btn event things.
        """
        if len(self._dev_common.find_watch()) == 0:
            raise serverNoResponse('RECIVE FAIL')

    def _generate_csv_file(self, name: str, data: dict):
        self._isAllPass = True
        folder = r'.\test_report\\'
        with open(folder + name + '.csv', 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows([
                ['Firmware version:', self._fw_ver],
                ['Firmware build time:', self._fw_build_t],
                [' '],
                ['Test case', 'Value/Error code', 'Note']
            ])
            for test_case, value in data.items():
                writer.writerow([test_case, value[0], value[1]])
                if value[1] != 'Pass':
                    self._isAllPass = False
    
    def _generate_xslx_file(self, name: str, data: dict):
        self._isAllPass = True
        wb = openpyxl.Workbook()
        s1 = wb[name]
        s1['A1'].value = 'Firmware version:'
        s1['A2'].value = self._fw_ver
        s1['B1'].value = 'Firmware build time:'
        s1['B2'].value = self._fw_build_t
        s1['D1'].value = 'Test case'
        s1['D2'].value = 'Value/Error code'
        s1['D3'].value = 'Note' 
        row = 4
        for test_case, value in data.items():
            row += 1
            s1.cell(row, 0).value = test_case
            s1.cell(row, 1).value = value[0]
            s1.cell(row, 2).value = value[1]
            if value[1] != 'Pass':
                self._isAllPass = False
                for column in range(3):
                    s1.cell(row, column).fill = PatternFill(
                        fill_type='solid', fgColor='FF0000'
                    )
        wb.save(name + '.xlsx')

    def do_export_log(self):
        """
        @brief: Do export file btn event things.
        """
        try:
            is_finish, is_already_read = False, False
            self._stop_bt_timer()
            if not self._dev_common.check_communicable():
                raise serverNoResponse('RECIVE FAIL')
            objTemp = log_readout()
            if not objTemp.is_read_out():
                sn = self._dev_common.get_sn()
                objTemp.start()
                data = objTemp.get_all_test_result()
                self._generate_csv_file(sn, data)
                is_finish, is_already_read = True, False
            else:
                is_finish, is_already_read = True, True
        finally:
            self._start_bt_timer()
            return is_finish, is_already_read
    
    def do_write_sn(self, sn: str):
        """
        @brief: Do write sn btn event things.
        """
        try:
            self._stop_bt_timer()
            if self._dev_common.check_communicable() == False:
                raise serverNoResponse('RECIVE FAIL')
            self._dev_common.set_sn(sn)
            res_sn = self._dev_common.get_sn()
            return (sn == res_sn, res_sn)
        finally:
            self._start_bt_timer()
    
    @do_in_thread
    def reflash_batt_volt(self):
        self._battery_volt = self._dev_common.get_battery_volt()

class bt_dev_display_info(object):
    def __init__(self, ble_addr: str, device_name: str, rssi: int, packet_data):
        self.addr = ble_addr
        self.device_name = device_name
        self.rssi = rssi
        self.packet_data = packet_data
        self.widgetItemText = f'{device_name} - ({rssi}) - {ble_addr}'
