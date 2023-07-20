#standard library
import sys
import os
import time
from enum import Enum
import csv
import serial
import re
from datetime import datetime
sys.path.append(os.getcwd())

#third party
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
import openpyxl
from openpyxl.styles import PatternFill

#module
from interface.interface_worker import  GenericWorker, do_in_thread
from interface.bluetooth.bluetooth import dongle
from common.convert import Data_convertor as cov
from common.config import Config_creator
from Script.Bluetooth.CR1_bt import test, device, fw_update
from Error.err_cls import *
from common.UI_object import Timer_generator
from page.factory_full_test.frontend import frontend
from page.factory_full_test.backend import backend
import logging
logger = logging.getLogger('factory_test')

#sub-module
if __name__ == '__main__':
    from UI_factory_test_page import Ui_factory_test_func
else:
    from .UI_factory_test_page import Ui_factory_test_func

cov = cov()
worker = GenericWorker()
do_in_thread = do_in_thread(worker)

class factory_test_func():
    def __init__(self):
        ## Init value ##
        self._is_connecting = False # flag for connecting, control Btn Behavior
        self.frontend = frontend()
        self.backend = backend()

        worker.message.connect(self.frontend.show_GUI_message)
        # self.LED_GuiNormal.set_color(2)
        self.Btn_scanDevice.clicked.connect(self._handle_scan_Btn_event)
        self.Btn_connect.clicked.connect(self.handle_connect_Btn_event)
        self.Btn_devTest.clicked.connect(self.handle_test_Btn_event)
        self.Btn_selectFile.clicked.connect(self._select_file)
        self.Btn_updataFw.clicked.connect(self.handle_updateFw_Btn_event)
        self.Btn_COMportOpen.clicked.connect(self.handle_COM_open_Btn_event)
        self.Btn_Calibrate.clicked.connect(self.handle_calPSensor_Btn_event)
        self.Btn_findWatch.clicked.connect(self._handle_find_watch_event)

    @do_in_thread
    def _handle_scan_Btn_event(self):
        try:
            self.frontend.set_scan_start()
            self.backend.do_scan()
            self.frontend.set_devicelist(self.backend.get_scan_data())
            self.frontend.set_scan_end()
        except Exception as e:
            self._handle_err(e)
    
    @do_in_thread
    def _connect_device(self):
        try:
            MAC_addrs = self.frontend.get_select_dev_MAC_addr()
            self.frontend.set_conn_start()
            self.backend.do_conn_dev(MAC_addrs)
        finally:
            self.frontend.set_conn_end(
                self._is_connecting,
                self.backend.get_fw_version(),
                self.backend.get_fw_build_time(),
                MAC_addrs
            )

    @do_in_thread
    def _disconnect_dev(self):
        self.frontend.set_disconn_start()
        self._is_connecting = self.backend.do_disconn_dev()
        self.frontend.set_disconn_end(self._is_connecting)

    @do_in_thread
    def handle_connect_Btn_event(self):
        try:
            self._disconnect_dev() if self._is_connecting else self._connect_device()
        except Exception as e:
            self._handle_err(e)

    @do_in_thread
    def handle_test_Btn_event(self):
        try:
            if self._is_connecting == False: 
                raise serverNoConnect()
            self.frontend.set_test_start()
            is_finish = self.backend.do_test()
        except Exception as e:
            self._handle_err(e)
        finally:
            data, is_all_pass = self.backend.get_test_data()
            self.frontend.set_test_end(is_finish, data, is_all_pass)

    def _generate_csv_file(self):
        self._isAllPass = True
        # folder = r'.\page\factory_test\test_report\\'
        folder = r'.\test_report\\'
        with open(folder + self._file_name + '.csv', 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows([
                ['Firmware version:', self._fw_version],
                ['Firmware build time:', self._fw_build_time],
                [' '],
                ['Test case', 'Value/Error code', 'Note']
            ])
            for test_case, value in self._file_data.items():
                writer.writerow([test_case, value[0], value[1]])
                if value[1] != 'Pass':
                    self._isAllPass = False
    
    def _generate_xslx_file(self):
        self._isAllPass = True
        wb = openpyxl.Workbook()
        s1 = wb[self._file_name]
        s1['A1'].value = 'Firmware version:'
        s1['A2'].value = self._fw_version
        s1['B1'].value = 'Firmware build time:'
        s1['B2'].value = self._fw_build_time
        s1['D1'].value = 'Test case'
        s1['D2'].value = 'Value/Error code'
        s1['D3'].value = 'Note' 
        row = 4
        for test_case, value in self._file_data.items():
            row += 1
            s1.cell(row, 0).value = test_case
            s1.cell(row, 1).value = value[0]
            s1.cell(row, 2).value = value[1]
            if value[1] != 'Pass':
                self._isAllPass = False
                for column in range(3):
                    s1.cell(row, column).fill = PatternFill(
                        fill_type='solid', fgColor='FF0000'
                    )
        wb.save(self._file_name + '.xlsx')

    def _select_file(self):
        filename, filetype = QFileDialog.getOpenFileName(self, "Open file", "./") # start path
        logger.info('Select file: ' + filename + '.' + filetype)
        worker.UI.emit(lambda: self.textEdit_filePath.setText(filename))
        self.textEdit_filePath.setFontFamily("Consolas")
        self.textEdit_filePath.setFontPointSize(16)

    @do_in_thread
    def handle_updateFw_Btn_event(self):
        try:
            if self._is_connecting == False:
                raise serverNoConnect()
            if self.textEdit_filePath.toPlainText() == '':
                raise conditionShort('NO FILE')
            self._set_Btn_disable(True)
            if not self._dev_common.check_communicable():
                raise serverNoResponse('TEST FAIL')

            # UI init
            self._show_GUI_message('START UPDATE FW')
            worker.UI.emit(lambda: self.Btn_updataFw.setText('UPDATING ...'))

            # action
            self._stop_bt_timer()
            self._dev_common.subscribe_ota_uuid()
            ota_update = fw_update(self.textEdit_filePath.toPlainText())
            error_code =  ota_update.start()
            if error_code is not None: # end of ota update
                self._show_GUI_message(str(error_code))
                self._disconnect_dev()
        except Exception as e:
            self._handle_err(e)
        finally: # UI deinit
            self._start_bt_timer()
            self._set_Btn_disable(False)
            worker.UI.emit(lambda: self.Btn_updataFw.setText('Finish Update'))
            time.sleep(1.5)
            worker.UI.emit(lambda: self.Btn_updataFw.setText('Update FW'))

    def handle_COM_open_Btn_event(self):
        try:
            self.frontend.set_open_com_start()
            port_num = self.frontend.get_com_port_num()
            self.backend.do_open_com_port(port_num)
        except IOError:
            logger.error('Cannot open COM' + port_num)
            self._show_GUI_message('COM PORT ERR')
        except Exception as e:
            self._handle_err(e)
        finally:
            self.frontend.set_open_com_end()

    @do_in_thread
    def handle_calPSensor_Btn_event(self):
        try:
            if not self._is_connecting:
                raise serverNoConnect()
            self.frontend.set_calPSensor_start()
            self.backend.do_cal_pSensor()
        except Exception as e:
            self._handle_err(e)
        finally: # UI deinit
            self.frontend.set_calPSensor_end()
    
    def _handle_err(self, err: Exception):
        err_type = type(err)
        if serverNoResponse == err_type:
            self._show_GUI_message(err.msg)
            self._disconnect_dev()
            self._set_Btn_disable(False)
        elif serverNoConnect == err_type:
            self._show_GUI_message('NO CONN')
        elif conditionShort == err_type:
            self._show_GUI_message(err.msg)
        else:
            # self.LED_GuiNormal.set_color(1)
            logger.error(err)
            self._show_GUI_message('UNKNOW ERR')

    @do_in_thread  
    def _set_Btn_disable(self, is_disable: bool):
        self.Btn_scanDevice.setDisabled(is_disable)
        self.Btn_connect.setDisabled(is_disable)
        self.Btn_devTest.setDisabled(is_disable)
        self.Btn_updataFw.setDisabled(is_disable)
        self.Btn_Calibrate.setDisabled(is_disable)
        self.Btn_findWatch.setDisabled(is_disable)
    
    def _show_GUI_message(self, info: str):
        self._gui_mes_sn += 1
        mbox_disp_inf = {
            'NO SELECT DEV': 'Please select a device, before connect.',
            'NO SN': 'Please enter the serial number of the device, before test begin.',
            'IN TESTING': 'The device is being tested, please wait for the test to finish and try again',
            'NO CONN': 'Please connect the device, and try again.',
            'START UPDATE FW': 'Start update firmware, please wait...',
            'FW UPDATE SUCC': 'FW update successfully.',
            'OPEN FILE ERR': 'Open file failed.',
            'FW HEADER ERR': 'Check firmware update file is effective.',
            'FW DATA ERR': 'Incomplete file, check firmware update file is effective.',
            'VERIFY FW ERR': 'firmware file has been sent, but the verifiy failed, please send the update file again',
            'COM PORT ERR': 'Cannot open COM port, please check the COM port is correct.',
            'COM PORT OK': 'COM port has been opened successly.',
            'TEST FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST1 FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST2 FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST3 FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST4 FAIL': 'Bluetooth device no response, please connect the device again.',
            'TEST5 FAIL': 'Bluetooth device no response, please connect the device again.',
            'SCAN START': 'Scanning for bleutooth devices, please wait...',
            'SCAN END': 'Scan finished.',
            'CONN START': 'Connecting to device, please wait...',
            'CONN END': 'Connect success.',
            'CONN FAIL': 'Connect failed, device may not trun on.',
            'CALIBRATE PASS': 'Calibrate barometric success.',
            'RECIVE FAIL': 'Bluetooth device no response',
            'Pressure Invalid': 'Instrument Pressure out of range',
            'TEST END': 'Test finished.',
            'DEV DISCONN': 'Bluetooth device disconnected.',
            'NO FILE': 'Please select a file, before update firmware.',
            'NO COM': 'Please open a COM port, and try again.',
            'UNKNOW ERR': 'Please turn off and remove the bluetooth receiver then plug it back in and turn on the GUI',
            'CALIBRATE FAIL': 'Calibrate barometric failed, device recive error.',
        }
        now = datetime.now()
        time = now.strftime("%Y-%m-%d %H:%M:%S")
        if info in mbox_disp_inf.keys():
            message = f'[{self._gui_mes_sn}]{time:<25}{info:<20}- {mbox_disp_inf[info]}'
        else:
            message = f'[{self._gui_mes_sn}]{time:<25}{info}'
        worker.UI.emit(lambda: self.textBrowser_guiMessage.append(message))

    @do_in_thread
    def _start_bt_timer(self):
        worker.UI.emit(lambda: self._timer['connChecker']['timer'].start(5000)) # ms
        worker.UI.emit(lambda: self._timer['awakener']['timer'].start(20000)) # ms
    
    @do_in_thread
    def _stop_bt_timer(self):
        worker.UI.emit(lambda: self._timer['connChecker']['timer'].stop())
        worker.UI.emit(lambda: self._timer['awakener']['timer'].stop())

    @do_in_thread
    def _handle_find_watch_event(self):
        try:
            if not self._is_connecting:
                raise serverNoConnect()
            if len(self._dev_common.find_watch()) == 0:
                raise serverNoResponse('TEST FAIL')
        except Exception as e:
            self._handle_err(e)
        

class bt_dev_display_info(object):
    def __init__(self, ble_addr: str, device_name: str, rssi: int, packet_data):
        self.addr = ble_addr
        self.device_name = device_name
        self.rssi = rssi
        self.packet_data = packet_data
        self.widgetItemText = f'{device_name} - ({rssi}) - {ble_addr}'

if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = factory_test_func()
    win.show()
    sys.exit(app.exec_())
